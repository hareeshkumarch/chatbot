import io

from openpyxl import load_workbook


def parse_xlsx(raw_bytes: bytes) -> list[tuple[str, int | None]]:
    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    blocks: list[str] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
        batch_size = 40
        for start in range(1, len(rows), batch_size):
            batch = rows[start:start + batch_size]
            lines = []
            for row in batch:
                pairs = [f"{header[i] if i < len(header) else f'col{i}'}: {value}" for i, value in enumerate(row)]
                lines.append(" | ".join(pairs))
            blocks.append(f"Sheet: {sheet.title}\n" + "\n".join(lines))
    return [(block, None) for block in blocks]
